"""
Google Play 评论 CSV → xlsx 预处理脚本

功能：读取一个或多个 Google Play 导出的评论 CSV，完成清洗、格式转换、拆分，
输出带机翻公式的 xlsx 文件（二级分类列留空，由 Claude 或 API 填充）。

用法：
    python review_csv_to_xlsx.py <input1.csv> [input2.csv ...] [-o output.xlsx]

依赖：
    pip install pandas openpyxl
"""

import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path


# ── 配置 ──────────────────────────────────────────────

SOURCE_COLUMNS = [
    "Review Text",
    "Star Rating",
    "Reviewer Language",
    "Review Submit Date and Time",
    "Review Link",
]

OUTPUT_HEADERS = [
    "二级分类",
    "机翻 (Translation)",
    "原文 (Review Text)",
    "评分 (Star Rating)",
    "语言代码 (Language)",
    "发布时间 (Submit Date)",
    "Review Link",
]

COLUMN_WIDTHS = [30, 45, 45, 12, 18, 18, 80]

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
BODY_FONT = Font(name="Arial", size=11)

FIVE_STAR_SHORT_THRESHOLD = 50  # 原文 < 此字符数的 5星评论自动标为纯好评


# ── Step 1: 读取 CSV ─────────────────────────────────

def read_csv(path: str) -> pd.DataFrame:
    """尝试 UTF-16，失败则回退 UTF-8。"""
    for enc in ("utf-16", "utf-8", "utf-8-sig", "cp1252"):
        try:
            df = pd.read_csv(path, encoding=enc)
            if all(col in df.columns for col in SOURCE_COLUMNS):
                return df[SOURCE_COLUMNS]
            missing = [c for c in SOURCE_COLUMNS if c not in df.columns]
            raise ValueError(f"CSV 缺少必需列: {missing}\n现有列: {list(df.columns)}")
        except (UnicodeDecodeError, UnicodeError):
            continue
    raise ValueError(f"无法以 UTF-16/UTF-8 解码文件: {path}")


# ── Step 2: 数据清洗 ─────────────────────────────────

def clean(df: pd.DataFrame) -> pd.DataFrame:
    df = df[df["Review Text"].notna() & (df["Review Text"].str.strip() != "")]
    df = df.copy()
    df["Review Submit Date and Time"] = (
        pd.to_datetime(df["Review Submit Date and Time"], errors="coerce")
        .dt.strftime("%Y-%m-%d")
    )
    return df


# ── Step 3: 拆分 ─────────────────────────────────────

def split_by_rating(df: pd.DataFrame):
    five_star = df[df["Star Rating"] == 5].reset_index(drop=True)
    low_star = df[df["Star Rating"] < 5].reset_index(drop=True)
    five_star_long = five_star[
        five_star["Review Text"].str.len() >= FIVE_STAR_SHORT_THRESHOLD
    ].reset_index(drop=True)
    return five_star_long, low_star


# ── Step 4 + 6: 写入 xlsx ────────────────────────────

def write_sheet(ws, df: pd.DataFrame):
    """向 worksheet 写入表头 + 数据行 + 机翻公式 + 样式。"""
    # 表头
    for col_idx, header in enumerate(OUTPUT_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 数据行
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        # A: 二级分类（留空）
        ws.cell(row=row_idx, column=1, value="")
        # B: 机翻公式
        formula = f'=IF(C{row_idx}="","",GOOGLETRANSLATE(C{row_idx},E{row_idx},"zh-CN"))'
        ws.cell(row=row_idx, column=2, value=formula)
        # C: 原文
        ws.cell(row=row_idx, column=3, value=row["Review Text"])
        # D: 评分
        ws.cell(row=row_idx, column=4, value=row["Star Rating"])
        # E: 语言代码
        ws.cell(row=row_idx, column=5, value=row["Reviewer Language"])
        # F: 发布时间
        ws.cell(row=row_idx, column=6, value=row["Review Submit Date and Time"])
        # G: Review Link
        ws.cell(row=row_idx, column=7, value=row["Review Link"])

    # 样式：列宽
    for col_idx, width in enumerate(COLUMN_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 字体
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(OUTPUT_HEADERS)):
        for cell in row:
            cell.font = BODY_FONT

    # 冻结首行 + 自动筛选
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:G{ws.max_row}"


def build_xlsx(all_low_star: pd.DataFrame, all_five_long: pd.DataFrame, output_path: str):
    wb = Workbook()

    # Sheet 1: 1-4星评论
    ws_low = wb.active
    ws_low.title = "1-4星评论"
    write_sheet(ws_low, all_low_star)

    # Sheet 2: 5星长评（≥50字符）
    ws_five_long = wb.create_sheet("5星长评")
    write_sheet(ws_five_long, all_five_long)

    wb.save(output_path)


# ── 主流程 ────────────────────────────────────────────

def main():
    # 解析参数
    args = sys.argv[1:]
    if not args:
        print("用法: python review_csv_to_xlsx.py <input1.csv> [input2.csv ...] [-o output.xlsx]")
        sys.exit(1)

    output_path = None
    input_paths = []
    i = 0
    while i < len(args):
        if args[i] == "-o" and i + 1 < len(args):
            output_path = args[i + 1]
            i += 2
        else:
            input_paths.append(args[i])
            i += 1

    if not input_paths:
        print("错误: 未指定输入 CSV 文件")
        sys.exit(1)

    if output_path is None:
        output_path = str(Path(input_paths[0]).with_suffix(".xlsx"))

    # 处理每个 CSV，合并结果
    all_low_star = []
    all_five_long = []
    total_raw = 0
    total_clean = 0

    for csv_path in input_paths:
        print(f"读取: {csv_path}")
        df = read_csv(csv_path)
        print(f"  原始行数: {len(df)}")
        total_raw += len(df)

        df = clean(df)
        print(f"  清洗后行数: {len(df)}")
        total_clean += len(df)

        five_star_long, low_star = split_by_rating(df)
        five_star_count = len(df[df["Star Rating"] == 5])
        five_star_short = five_star_count - len(five_star_long)
        print(f"  1-4星: {len(low_star)} 条, 5星长评: {len(five_star_long)} 条, 5星短评(跳过): {five_star_short} 条")

        all_low_star.append(low_star)
        all_five_long.append(five_star_long)

    # 合并
    all_low_star = pd.concat(all_low_star, ignore_index=True) if all_low_star else pd.DataFrame()
    all_five_long = pd.concat(all_five_long, ignore_index=True) if all_five_long else pd.DataFrame()

    build_xlsx(all_low_star, all_five_long, output_path)

    print(f"\n汇总: {len(input_paths)} 个 CSV, 原始 {total_raw} 条, 清洗后 {total_clean} 条")
    print(f"输出: {output_path}")
    print(f"  ✓ Sheet 1: 1-4星评论 ({len(all_low_star)} 条)")
    print(f"  ✓ Sheet 2: 5星长评 ({len(all_five_long)} 条)")
    print("  ✓ 机翻公式已写入，上传 Google Sheets 后生效")


if __name__ == "__main__":
    main()
